"""
Exporte .data/birthdays.xlsx → .data/birthdays.csv

Workflow :
  1. Modifier .data/birthdays.xlsx dans Excel
     (colonnes : name, month, day, birth_year, display)
  2. Lancer : python gen_birthdays.py
  3. Committer les changements (birthdays.csv sera mis à jour)

Colonne display : 1 = affiché dans le dashboard, 0 = masqué
"""
import csv, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl requis : pip install openpyxl")

data_dir  = Path(__file__).parent / '.data'
xlsx_path = data_dir / 'birthdays.xlsx'
csv_path  = data_dir / 'birthdays.csv'

if not xlsx_path.exists():
    sys.exit(f"Fichier introuvable : {xlsx_path}")

wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
ws = wb.active

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
# Colonnes attendues (insensible à la casse)
h_lower = [str(h).lower().strip() if h else '' for h in headers]

def col(name):
    return h_lower.index(name)

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    try:
        name       = str(row[col('name')]).strip()
        month      = int(row[col('month')])
        day        = int(row[col('day')])
        birth_year = row[col('birth_year')]
        display    = int(row[col('display')]) if 'display' in h_lower else 1
    except (ValueError, TypeError, IndexError):
        continue
    if not name or not 1 <= month <= 12 or not 1 <= day <= 31:
        continue
    birth_year_str = str(int(birth_year)) if birth_year not in (None, '') else ''
    rows.append([name, month, day, birth_year_str, display])

# Tri : mois, jour, nom
rows.sort(key=lambda x: (x[1], x[2], x[0]))

with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(['name', 'month', 'day', 'birth_year', 'display'])
    writer.writerows(rows)

print(f"{len(rows)} anniversaires exportes -> {csv_path}")
